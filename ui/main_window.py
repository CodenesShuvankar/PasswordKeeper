"""
Main application window
"""

import os
import logging
import csv
from datetime import datetime
from typing import Optional, List, Dict, Any
from PyQt6.QtWidgets import (
    QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QTableWidget, 
    QTableWidgetItem, QPushButton, QLineEdit, QComboBox, QLabel,
    QMessageBox, QDialog, QMenu, QMenuBar, QStatusBar, QSplitter,
    QGroupBox, QFormLayout, QTextEdit, QHeaderView, QAbstractItemView,
    QInputDialog, QCheckBox, QSpinBox, QTabWidget, QFileDialog
)
from PyQt6.QtCore import Qt, QTimer, pyqtSignal, QThread, pyqtSlot
from PyQt6.QtGui import QAction, QIcon, QFont, QPixmap, QKeySequence, QColor

try:
    from openpyxl import Workbook, load_workbook
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

from core.db import DatabaseManager
from core.auth import AuthManager
from core.utils import ClipboardManager, InactivityTimer, SettingsManager
from .credential_dialog import CredentialDialog
from .credential_view_dialog import CredentialViewDialog
from .settings_dialog import SettingsDialog
from .password_generator import PasswordGeneratorDialog


class MainWindow(QMainWindow):
    """Main application window with modern UI design"""
    
    def __init__(self, db_manager: DatabaseManager, auth_manager: AuthManager):
        super().__init__()
        
        self.db_manager = db_manager
        self.auth_manager = auth_manager
        self.logger = logging.getLogger(__name__)
        
        # Managers
        self.settings_manager = SettingsManager()
        self.clipboard_manager = ClipboardManager(
            self.settings_manager.get('clipboard_auto_clear_seconds', 30)
        )
        self.inactivity_timer = InactivityTimer(
            self.settings_manager.get('inactivity_timeout_minutes', 5)
        )
        
        # Data
        self.credentials = []
        self.filtered_credentials = []
        self.current_category = "All"
        
        # Apply modern styling
        self.apply_modern_style()
        
        # UI Components
        self.central_widget = None
        self.table_widget = None
        self.search_line_edit = None
        self.category_combo = None
        self.count_label = None
        self.status_bar = None
        
        self.init_ui()
        self.setup_menus()
        self.setup_status_bar()
        self.setup_shortcuts()
        self.load_data()
        self.apply_settings()
        
        # Setup inactivity timer
        self.inactivity_timer.start(self.auto_lock)
        
        # Auto-save timer
        self.auto_save_timer = QTimer()
        self.auto_save_timer.timeout.connect(self.auto_save)
        self.auto_save_timer.start(30000)  # Auto-save every 30 seconds
        
        self.logger.info("Main window initialized")
        # Disable dynamic font scaling to avoid CSS parsing issues
        # self._base_width_for_scale = 1300
        # self._min_scale = 0.85
        # self._max_scale = 1.25
        # self.apply_dynamic_text_size()

    # ---------------- Dynamic Text Size Logic ---------------- #
    def compute_scale_factor(self) -> float:
        """Compute scale factor based on current window width.
        Base width = 1300 (design width). Clamped to [0.85, 1.25]."""
        w = max(400, self.width())
        raw = w / float(self._base_width_for_scale)
        return max(self._min_scale, min(self._max_scale, raw))

    def apply_dynamic_text_size(self):
        """Apply dynamic font sizes to key UI elements (header title + credential table)."""
        try:
            scale = self.compute_scale_factor()

            # Scale header title if available
            if hasattr(self, 'page_title') and self.page_title is not None:
                f = self.page_title.font()
                f.setPointSize(int(24 * scale))
                self.page_title.setFont(f)

            # Scale count label
            if hasattr(self, 'count_label') and self.count_label is not None:
                cf = self.count_label.font()
                cf.setPointSize(int(12 * scale))
                self.count_label.setFont(cf)

            # Scale table fonts via style regeneration
            if hasattr(self, 'table_widget') and self.table_widget is not None and hasattr(self, '_table_style_template'):
                body_size = 11.5 * scale
                header_size = 11 * scale
                pad_v = int(18 * scale)
                pad_h = int(14 * scale)
                style = self._table_style_template.format(
                    TABLE_FONT_SIZE=f"{body_size:.2f}pt",
                    HEADER_FONT_SIZE=f"{header_size:.2f}pt",
                    ITEM_PADDING_V=pad_v,
                    ITEM_PADDING_H=pad_h
                )
                self.table_widget.setStyleSheet(style)
        except Exception as e:
            # Log error but don't crash the application
            self.logger.error(f"Error applying dynamic text size: {e}")
            # Fall back to static style if dynamic scaling fails
            if hasattr(self, 'table_widget') and self.table_widget is not None:
                static_style = """
                    QTableWidget {
                        background: white;
                        border: 2px solid rgba(31, 47, 155, 0.1);
                        border-radius: 15px;
                        gridline-color: rgba(31, 47, 155, 0.1);
                        selection-background-color: rgba(123, 213, 245, 0.2);
                        color: #1F2F9B;
                        font-size: 11.5pt;
                    }
                    QTableWidget::item {
                        padding: 18px 14px;
                        border-bottom: 1px solid rgba(31, 47, 155, 0.1);
                        border-right: 1px solid rgba(31, 47, 155, 0.05);
                    }
                    QTableWidget::item:selected {
                        background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                            stop:0 rgba(123, 213, 245, 0.3), stop:1 rgba(76, 222, 222, 0.3));
                        color: #1F2F9B;
                        font-weight: bold;
                    }
                    QHeaderView::section {
                        background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                            stop:0 #1F2F9B, stop:1 #1CA7EC);
                        color: white;
                        padding: 15px 10px;
                        border: none;
                        border-right: 1px solid rgba(255, 255, 255, 0.2);
                        font-weight: bold;
                        font-size: 11pt;
                    }
                    QHeaderView::section:first { border-top-left-radius: 12px; }
                    QHeaderView::section:last { border-top-right-radius: 12px; }
                """
                self.table_widget.setStyleSheet(static_style)

    def resizeEvent(self, event):  # type: ignore[override]
        super().resizeEvent(event)
        # Disabled dynamic text scaling to avoid CSS issues
        # self.apply_dynamic_text_size()
    
    def apply_modern_style(self):
        """Apply modern UI styling with dual-shade color scheme"""
        style = """
        /* Main Window */
        QMainWindow {
            background: qlineargradient(x1:0, y1:0, x2:1, y2:1, 
                stop:0 #1F2F9B, stop:1 #1CA7EC);
            color: white;
            font-family: 'Segoe UI', Arial, sans-serif;
        }
        
        /* Central Widget */
        QWidget {
            background-color: transparent;
            color: white;
            font-size: 11pt;
        }
        
        /* Card-style containers with responsive design */
        QGroupBox {
            background: qlineargradient(x1:0, y1:0, x2:1, y2:1, 
                stop:0 rgba(123, 213, 245, 0.2), stop:1 rgba(76, 222, 222, 0.2));
            border: 2px solid rgba(255, 255, 255, 0.3);
            border-radius: 15px;
            margin: 5px;
            padding: 10px;
            font-weight: bold;
            font-size: 12pt;
            min-height: 40px;
        }
        
        QGroupBox::title {
            color: #7BD5F5;
            subcontrol-origin: margin;
            left: 20px;
            padding: 5px 15px;
            background: rgba(31, 47, 155, 0.8);
            border-radius: 8px;
            border: 1px solid rgba(255, 255, 255, 0.3);
        }
        
        /* Responsive Modern Buttons */
        QPushButton {
            background: qlineargradient(x1:0, y1:0, x2:1, y2:1, 
                stop:0 #7BD5F5, stop:1 #4ADEDE);
            color: #1F2F9B;
            border: none;
            border-radius: 18px;
            padding: 8px 16px;
            font-weight: bold;
            font-size: 10pt;
            min-height: 16px;
            min-width: 60px;
        }
        
        QPushButton:hover {
            background: qlineargradient(x1:0, y1:0, x2:1, y2:1, 
                stop:0 #4ADEDE, stop:1 #7BD5F5);
            transform: translateY(-2px);
            box-shadow: 0 4px 15px rgba(123, 213, 245, 0.4);
        }
        
        QPushButton:pressed {
            background: qlineargradient(x1:0, y1:0, x2:1, y2:1, 
                stop:0 #1CA7EC, stop:1 #1F2F9B);
            color: white;
            transform: translateY(0px);
        }
        
        QPushButton:disabled {
            background: rgba(255, 255, 255, 0.1);
            color: rgba(255, 255, 255, 0.4);
        }
        
        /* Responsive Primary Action Buttons */
        QPushButton#primaryButton {
            background: qlineargradient(x1:0, y1:0, x2:1, y2:1, 
                stop:0 #1CA7EC, stop:1 #1F2F9B);
            color: white;
            border-radius: 20px;
            padding: 10px 20px;
            font-size: 11pt;
            font-weight: bold;
            min-width: 100px;
            min-height: 35px;
        }
        
        QPushButton#primaryButton:hover {
            background: qlineargradient(x1:0, y1:0, x2:1, y2:1, 
                stop:0 #7BD5F5, stop:1 #4ADEDE);
            color: #1F2F9B;
        }
        
        /* Responsive Floating Action Buttons */
        QPushButton#floatingButton {
            background: qlineargradient(x1:0, y1:0, x2:1, y2:1, 
                stop:0 #7BD5F5, stop:1 #4ADEDE);
            color: #1F2F9B;
            border: 2px solid rgba(255, 255, 255, 0.5);
            border-radius: 25px;
            padding: 10px;
            font-weight: bold;
            font-size: 14pt;
            min-width: 50px;
            min-height: 50px;
            max-width: 50px;
            max-height: 50px;
        }
        }
        
        QPushButton#floatingButton:hover {
            background: qlineargradient(x1:0, y1:0, x2:1, y2:1, 
                stop:0 #4ADEDE, stop:1 #7BD5F5);
            border: 3px solid white;
            transform: scale(1.1);
        }
        
        /* Modern Table */
        QTableWidget {
            background: rgba(255, 255, 255, 0.1);
            border: 2px solid rgba(255, 255, 255, 0.3);
            border-radius: 15px;
            selection-background-color: rgba(123, 213, 245, 0.3);
            gridline-color: rgba(255, 255, 255, 0.2);
            color: white;
            font-size: 11pt;
        }
        
        QTableWidget::item {
            padding: 10px;
            border-bottom: 1px solid rgba(255, 255, 255, 0.1);
        }
        
        QTableWidget::item:selected {
            background: qlineargradient(x1:0, y1:0, x2:1, y2:1, 
                stop:0 rgba(123, 213, 245, 0.4), stop:1 rgba(76, 222, 222, 0.4));
            color: white;
        }
        
        QHeaderView::section {
            background: qlineargradient(x1:0, y1:0, x2:1, y2:1, 
                stop:0 #1F2F9B, stop:1 #1CA7EC);
            color: white;
            padding: 12px;
            border: none;
            border-right: 1px solid rgba(255, 255, 255, 0.2);
            font-weight: bold;
            font-size: 11pt;
        }
        
        QHeaderView::section:first {
            border-top-left-radius: 10px;
        }
        
        QHeaderView::section:last {
            border-top-right-radius: 10px;
        }
        
        /* Modern Input Fields */
        QLineEdit, QComboBox {
            background: rgba(255, 255, 255, 0.15);
            border: 2px solid rgba(255, 255, 255, 0.3);
            border-radius: 12px;
            padding: 10px 15px;
            color: white;
            font-size: 11pt;
            selection-background-color: rgba(123, 213, 245, 0.5);
        }
        
        QLineEdit:focus, QComboBox:focus {
            border: 2px solid #7BD5F5;
            background: rgba(255, 255, 255, 0.2);
        }
        
        QComboBox::drop-down {
            border: none;
            width: 30px;
        }
        
        QComboBox::down-arrow {
            image: none;
            border-style: solid;
            border-width: 6px 6px 0 6px;
            border-color: #7BD5F5 transparent transparent transparent;
        }
        
        QComboBox QAbstractItemView {
            background: rgba(31, 47, 155, 0.95);
            border: 2px solid #7BD5F5;
            border-radius: 10px;
            selection-background-color: rgba(123, 213, 245, 0.3);
            color: white;
        }
        
        /* Status Bar */
        QStatusBar {
            background: qlineargradient(x1:0, y1:0, x2:1, y2:0, 
                stop:0 rgba(31, 47, 155, 0.8), stop:1 rgba(28, 167, 236, 0.8));
            color: white;
            border-top: 2px solid rgba(255, 255, 255, 0.3);
            padding: 5px;
        }
        
        /* Labels */
        QLabel {
            color: white;
            font-size: 11pt;
        }
        
        /* Menu Bar */
        QMenuBar {
            background: qlineargradient(x1:0, y1:0, x2:1, y2:0, 
                stop:0 #1F2F9B, stop:1 #1CA7EC);
            color: white;
            border-bottom: 2px solid rgba(255, 255, 255, 0.3);
            padding: 5px;
        }
        
        QMenuBar::item {
            background: transparent;
            padding: 8px 15px;
            border-radius: 8px;
        }
        
        QMenuBar::item:selected {
            background: rgba(123, 213, 245, 0.3);
        }
        
        QMenu {
            background: rgba(31, 47, 155, 0.95);
            border: 2px solid #7BD5F5;
            border-radius: 10px;
            color: white;
        }
        
        QMenu::item {
            padding: 10px 20px;
        }
        
        QMenu::item:selected {
            background: rgba(123, 213, 245, 0.3);
        }
        
        /* Scrollbars */
        QScrollBar:vertical {
            background: rgba(255, 255, 255, 0.1);
            width: 12px;
            border-radius: 6px;
        }
        
        QScrollBar::handle:vertical {
            background: qlineargradient(x1:0, y1:0, x2:1, y2:0, 
                stop:0 #7BD5F5, stop:1 #4ADEDE);
            border-radius: 6px;
            min-height: 20px;
        }
        
        QScrollBar::handle:vertical:hover {
            background: qlineargradient(x1:0, y1:0, x2:1, y2:0, 
                stop:0 #4ADEDE, stop:1 #7BD5F5);
        }
        """
        
        self.setStyleSheet(style)
    
    def init_ui(self):
        """Initialize the modern sidebar-based user interface"""
        self.setWindowTitle("🔐 Password Keeper - Secure & Modern")
        self.setMinimumSize(1000, 700)
        self.resize(1300, 900)  # Larger default size for sidebar layout
        
        # Central widget
        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)
        
        # Main horizontal layout (sidebar + content)
        main_layout = QHBoxLayout(self.central_widget)
        main_layout.setSpacing(0)
        main_layout.setContentsMargins(0, 0, 0, 0)
        
        # Create sidebar
        self.sidebar = self.create_sidebar()
        main_layout.addWidget(self.sidebar, 0)  # Fixed width
        
        # Create main content area
        self.content_area = self.create_content_area()
        main_layout.addWidget(self.content_area, 1)  # Takes remaining space
    
    def create_sidebar(self) -> QWidget:
        """Create modern sidebar with navigation and actions"""
        sidebar = QWidget()
        sidebar.setFixedWidth(280)
        sidebar.setStyleSheet("""
            QWidget {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1, 
                    stop:0 rgba(31, 47, 155, 0.95), stop:1 rgba(28, 167, 236, 0.95));
                border-right: 3px solid rgba(255, 255, 255, 0.2);
            }
        """)
        
        layout = QVBoxLayout(sidebar)
        layout.setSpacing(20)
        layout.setContentsMargins(20, 30, 20, 30)
        
        # App title/logo
        title_label = QLabel("🔐 Password Keeper")
        title_label.setStyleSheet("""
            QLabel {
                color: #7BD5F5;
                font-size: 18pt;
                font-weight: bold;
                padding: 15px;
                background: rgba(255, 255, 255, 0.1);
                border-radius: 15px;
                border: 2px solid rgba(255, 255, 255, 0.2);
            }
        """)
        title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(title_label)
        
        layout.addSpacing(20)
        
        # Navigation section
        nav_label = QLabel("📚 MANAGEMENT")
        nav_label.setStyleSheet("""
            QLabel {
                color: rgba(255, 255, 255, 0.7);
                font-size: 10pt;
                font-weight: bold;
                padding: 5px 0px;
            }
        """)
        layout.addWidget(nav_label)
        
        # Credentials management button
        self.credentials_nav_btn = QPushButton("🗂️  CREDENTIALS")
        self.credentials_nav_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0, 
                    stop:0 rgba(123, 213, 245, 0.3), stop:1 rgba(76, 222, 222, 0.3));
                color: white;
                border: 2px solid rgba(255, 255, 255, 0.3);
                border-radius: 12px;
                padding: 15px;
                font-size: 12pt;
                font-weight: bold;
                text-align: left;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0, 
                    stop:0 rgba(123, 213, 245, 0.5), stop:1 rgba(76, 222, 222, 0.5));
                border: 2px solid rgba(255, 255, 255, 0.5);
            }
        """)
        layout.addWidget(self.credentials_nav_btn)
        
        layout.addSpacing(20)
        
        # Actions section
        actions_label = QLabel("⚡ ACTIONS")
        actions_label.setStyleSheet("""
            QLabel {
                color: rgba(255, 255, 255, 0.7);
                font-size: 10pt;
                font-weight: bold;
                padding: 5px 0px;
            }
        """)
        layout.addWidget(actions_label)
        
        # Action buttons
        self.add_btn = QPushButton("➕  Add Credential")
        self.password_gen_btn = QPushButton("🔑  Generate Password")
        self.import_btn = QPushButton("📥  Import Data")
        self.export_btn = QPushButton("📤  Export Data")
        self.backup_btn = QPushButton("💾  Create Backup")
        
        sidebar_buttons = [
            self.add_btn, self.password_gen_btn, self.import_btn, 
            self.export_btn, self.backup_btn
        ]
        
        for btn in sidebar_buttons:
            btn.setStyleSheet("""
                QPushButton {
                    background: rgba(255, 255, 255, 0.1);
                    color: white;
                    border: 1px solid rgba(255, 255, 255, 0.2);
                    border-radius: 10px;
                    padding: 12px;
                    font-size: 11pt;
                    font-weight: bold;
                    text-align: left;
                }
                QPushButton:hover {
                    background: qlineargradient(x1:0, y1:0, x2:1, y2:0, 
                        stop:0 #7BD5F5, stop:1 #4ADEDE);
                    color: #1F2F9B;
                    border: 1px solid rgba(255, 255, 255, 0.5);
                }
                QPushButton:pressed {
                    background: qlineargradient(x1:0, y1:0, x2:1, y2:0, 
                        stop:0 #1CA7EC, stop:1 #1F2F9B);
                    color: white;
                }
            """)
            layout.addWidget(btn)
        
        # Connect signals
        self.add_btn.clicked.connect(self.add_credential)
        self.password_gen_btn.clicked.connect(self.show_password_generator)
        self.import_btn.clicked.connect(self.import_from_excel)
        self.export_btn.clicked.connect(self.export_to_excel)
        
        layout.addStretch()
        
        # Settings section at bottom
        settings_label = QLabel("⚙️ SETTINGS")
        settings_label.setStyleSheet("""
            QLabel {
                color: rgba(255, 255, 255, 0.7);
                font-size: 10pt;
                font-weight: bold;
                padding: 5px 0px;
            }
        """)
        layout.addWidget(settings_label)
        
        self.settings_btn = QPushButton("🔧  Settings")
        self.lock_btn = QPushButton("🔒  Lock App")
        
        for btn in [self.settings_btn, self.lock_btn]:
            btn.setStyleSheet("""
                QPushButton {
                    background: rgba(255, 255, 255, 0.1);
                    color: white;
                    border: 1px solid rgba(255, 255, 255, 0.2);
                    border-radius: 10px;
                    padding: 12px;
                    font-size: 11pt;
                    font-weight: bold;
                    text-align: left;
                }
                QPushButton:hover {
                    background: qlineargradient(x1:0, y1:0, x2:1, y2:0, 
                        stop:0 #7BD5F5, stop:1 #4ADEDE);
                    color: #1F2F9B;
                }
            """)
            layout.addWidget(btn)
        
        self.lock_btn.clicked.connect(self.lock_application)
        
        return sidebar
    
    def create_content_area(self) -> QWidget:
        """Create main content area with header and credential list"""
        content = QWidget()
        content.setStyleSheet("""
            QWidget {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1, 
                    stop:0 rgba(240, 248, 255, 0.95), stop:1 rgba(230, 245, 255, 0.95));
            }
        """)
        
        layout = QVBoxLayout(content)
        layout.setSpacing(20)
        layout.setContentsMargins(30, 30, 30, 30)
        
        # Header section
        header_layout = QHBoxLayout()
        
        # Page title
        page_title = QLabel("Credential Management")
        page_title.setStyleSheet("""
            QLabel {
                color: #1F2F9B;
                /* base size; dynamically overridden */
                font-size: 24pt;
                font-weight: bold;
                padding: 10px 0px;
            }
        """)
        # Store reference for dynamic scaling
        self.page_title = page_title
        header_layout.addWidget(page_title)
        
        header_layout.addStretch()
        
        # Header action buttons
        self.view_btn = QPushButton("👁️ View")
        self.edit_btn = QPushButton("✏️ Edit")
        self.delete_btn = QPushButton("🗑️ Delete")
        
        header_buttons = [self.view_btn, self.edit_btn, self.delete_btn]
        for btn in header_buttons:
            btn.setEnabled(False)
            btn.setStyleSheet("""
                QPushButton {
                    background: qlineargradient(x1:0, y1:0, x2:1, y2:1, 
                        stop:0 #7BD5F5, stop:1 #4ADEDE);
                    color: #1F2F9B;
                    border: none;
                    border-radius: 8px;
                    padding: 10px 20px;
                    font-weight: bold;
                    font-size: 11pt;
                    margin: 0px 5px;
                }
                QPushButton:hover {
                    background: qlineargradient(x1:0, y1:0, x2:1, y2:1, 
                        stop:0 #4ADEDE, stop:1 #7BD5F5);
                }
                QPushButton:disabled {
                    background: rgba(200, 200, 200, 0.3);
                    color: rgba(100, 100, 100, 0.7);
                }
            """)
            header_layout.addWidget(btn)
        
        # Connect header button signals
        self.view_btn.clicked.connect(self.view_credential_details)
        self.edit_btn.clicked.connect(self.edit_credential)
        self.delete_btn.clicked.connect(self.delete_credential)
        
        layout.addLayout(header_layout)
        
        # Search and filter bar
        filter_layout = self.create_modern_filter_bar()
        layout.addLayout(filter_layout)
        
        # Credentials list view
        self.create_modern_credentials_list()
        layout.addWidget(self.table_widget, 1)  # Takes most space
        
        # Status bar
        status_layout = QHBoxLayout()
        self.count_label = QLabel("0 credentials")
        self.count_label.setStyleSheet("""
            QLabel {
                color: #1F2F9B;
                font-size: 12pt;
                font-weight: bold;
                padding: 10px;
            }
        """)
        status_layout.addWidget(self.count_label)
        status_layout.addStretch()
        
        # Quick action buttons in content area
        self.copy_username_btn = QPushButton("👤 Copy Username")
        self.copy_password_btn = QPushButton("🔑 Copy Password")
        self.open_url_btn = QPushButton("🌐 Open URL")
        
        quick_buttons = [self.copy_username_btn, self.copy_password_btn, self.open_url_btn]
        for btn in quick_buttons:
            btn.setEnabled(False)
            btn.setStyleSheet("""
                QPushButton {
                    background: rgba(31, 47, 155, 0.1);
                    color: #1F2F9B;
                    border: 1px solid rgba(31, 47, 155, 0.3);
                    border-radius: 8px;
                    padding: 8px 16px;
                    font-weight: bold;
                    font-size: 10pt;
                    margin: 0px 3px;
                }
                QPushButton:hover {
                    background: qlineargradient(x1:0, y1:0, x2:1, y2:1, 
                        stop:0 #7BD5F5, stop:1 #4ADEDE);
                    color: #1F2F9B;
                }
                QPushButton:disabled {
                    background: rgba(200, 200, 200, 0.2);
                    color: rgba(100, 100, 100, 0.5);
                }
            """)
            status_layout.addWidget(btn)
        
        # Connect quick action signals
        self.copy_username_btn.clicked.connect(self.copy_username)
        self.copy_password_btn.clicked.connect(self.copy_password)
        self.open_url_btn.clicked.connect(self.open_url)
        
        layout.addLayout(status_layout)
        
        return content
    
    def create_modern_filter_bar(self) -> QHBoxLayout:
        """Create modern search and filter bar for content area"""
        layout = QHBoxLayout()
        layout.setSpacing(15)
        
        # Search container
        search_container = QWidget()
        search_container.setStyleSheet("""
            QWidget {
                background: white;
                border: 2px solid rgba(31, 47, 155, 0.2);
                border-radius: 12px;
                padding: 5px;
            }
        """)
        search_layout = QHBoxLayout(search_container)
        search_layout.setContentsMargins(15, 8, 15, 8)
        search_layout.setSpacing(10)
        
        search_icon = QLabel("🔍")
        search_icon.setStyleSheet("color: #1F2F9B; font-size: 14pt;")
        search_layout.addWidget(search_icon)
        
        self.search_line_edit = QLineEdit()
        self.search_line_edit.setPlaceholderText("Search credentials...")
        self.search_line_edit.setStyleSheet("""
            QLineEdit {
                border: none;
                background: transparent;
                color: #1F2F9B;
                font-size: 12pt;
                padding: 5px;
            }
        """)
        self.search_line_edit.textChanged.connect(self.filter_credentials)
        search_layout.addWidget(self.search_line_edit, 1)
        
        layout.addWidget(search_container, 2)
        
        # Category filter
        category_container = QWidget()
        category_container.setStyleSheet("""
            QWidget {
                background: white;
                border: 2px solid rgba(31, 47, 155, 0.2);
                border-radius: 12px;
                padding: 5px;
            }
        """)
        category_layout = QHBoxLayout(category_container)
        category_layout.setContentsMargins(15, 8, 15, 8)
        category_layout.setSpacing(10)
        
        category_icon = QLabel("📁")
        category_icon.setStyleSheet("color: #1F2F9B; font-size: 14pt;")
        category_layout.addWidget(category_icon)
        
        self.category_combo = QComboBox()
        self.category_combo.setStyleSheet("""
            QComboBox {
                border: 1px solid rgba(31, 47, 155, 0.3);
                background: white;
                color: #1F2F9B;
                font-size: 12pt;
                padding: 5px 30px 5px 10px;
                min-width: 120px;
                min-height: 30px;
                border-radius: 6px;
            }
            QComboBox:hover {
                border: 1px solid #1CA7EC;
                background: rgba(123, 213, 245, 0.1);
            }
            QComboBox:focus {
                border: 2px solid #1CA7EC;
                outline: none;
            }
            QComboBox::drop-down {
                subcontrol-origin: padding;
                subcontrol-position: top right;
                width: 25px;
                border: none;
                background: transparent;
            }
            QComboBox::drop-down:hover {
                background: rgba(123, 213, 245, 0.2);
                border-radius: 4px;
            }
            QComboBox::down-arrow {
                image: none;
                border-style: solid;
                border-width: 8px 8px 0 8px;
                border-color: #1F2F9B transparent transparent transparent;
                width: 0px;
                height: 0px;
                margin-right: 5px;
            }
            QComboBox::down-arrow:hover {
                border-color: #1CA7EC transparent transparent transparent;
            }
            QComboBox QAbstractItemView {
                background: white;
                border: 2px solid #7BD5F5;
                border-radius: 8px;
                selection-background-color: rgba(123, 213, 245, 0.3);
                color: #1F2F9B;
                font-size: 11pt;
                padding: 4px;
                min-width: 140px;
                outline: none;
            }
            QComboBox QAbstractItemView::item {
                padding: 10px 12px;
                border: none;
                min-height: 25px;
            }
            QComboBox QAbstractItemView::item:hover {
                background-color: rgba(123, 213, 245, 0.2);
            }
            QComboBox QAbstractItemView::item:selected {
                background-color: rgba(123, 213, 245, 0.4);
                color: #1F2F9B;
                font-weight: bold;
            }
        """)
        self.category_combo.currentTextChanged.connect(self.filter_by_category)
        
        # Initialize with default categories
        default_categories = ["All", "Email", "Shopping", "Social Media", "Banking", "Work", "Personal", "Other"]
        self.category_combo.addItems(default_categories)
        
        category_layout.addWidget(self.category_combo, 1)
        
        layout.addWidget(category_container, 1)
        
        return layout
    
    def create_modern_credentials_list(self):
        """Create modern list-style credentials table (avatar column removed)"""
        self.table_widget = QTableWidget()
        # Now 5 columns: Title, Username, URL, Category, Modified
        self.table_widget.setColumnCount(5)
        self.table_widget.setHorizontalHeaderLabels([
            "Title", "Username", "URL", "Category", "Modified"
        ])
        
        # Modern table styling (static)
        self.table_widget.setStyleSheet("""
            QTableWidget {
                background: white;
                border: 2px solid rgba(31, 47, 155, 0.1);
                border-radius: 15px;
                gridline-color: rgba(31, 47, 155, 0.1);
                selection-background-color: rgba(123, 213, 245, 0.2);
                color: #1F2F9B;
                font-size: 11.5pt;
            }
            QTableWidget::item {
                padding: 18px 14px;
                border-bottom: 1px solid rgba(31, 47, 155, 0.1);
                border-right: 1px solid rgba(31, 47, 155, 0.05);
            }
            QTableWidget::item:selected {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 rgba(123, 213, 245, 0.3), stop:1 rgba(76, 222, 222, 0.3));
                color: #1F2F9B;
                font-weight: bold;
            }
            QHeaderView::section {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 #1F2F9B, stop:1 #1CA7EC);
                color: white;
                padding: 15px 10px;
                border: none;
                border-right: 1px solid rgba(255, 255, 255, 0.2);
                font-weight: bold;
                font-size: 11pt;
            }
            QHeaderView::section:first { border-top-left-radius: 12px; }
            QHeaderView::section:last { border-top-right-radius: 12px; }
        """)
        
        # Remove template and dynamic application
        # self._table_style_template = """..."""  
        # self.apply_dynamic_text_size()
        
        # Table settings for modern list view
        self.table_widget.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table_widget.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.table_widget.setAlternatingRowColors(False)  # We'll handle this with CSS
        self.table_widget.setSortingEnabled(True)
        self.table_widget.setShowGrid(True)
        
        # Modern column sizing like the reference image
        header = self.table_widget.horizontalHeader()
        header.setStretchLastSection(False)
        
        # Set column widths (redistribute space previously used by avatar)
        self.table_widget.setColumnWidth(0, 230)  # Title - slightly larger
        self.table_widget.setColumnWidth(1, 170)  # Username
        self.table_widget.setColumnWidth(2, 300)  # URL - large
        self.table_widget.setColumnWidth(3, 140)  # Category
        self.table_widget.setColumnWidth(4, 160)  # Modified

        # Resize behavior: stretch main text columns, allow manual for others
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)     # Title
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.Interactive) # Username
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)     # URL
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.Interactive) # Category
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.Interactive) # Modified
        
        # Set row height for list-like appearance
        self.table_widget.verticalHeader().setDefaultSectionSize(66)  # Slightly taller for new padding
        self.table_widget.verticalHeader().setVisible(False)
        
        # Connect signals
        self.table_widget.itemSelectionChanged.connect(self.on_selection_changed)
        self.table_widget.itemDoubleClicked.connect(self.view_credential_details)
        
        # Context menu
        self.table_widget.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.table_widget.customContextMenuRequested.connect(self.show_context_menu)
    
    def setup_menus(self):
        """Setup application menus"""
        menubar = self.menuBar()
        
        # File menu
        file_menu = menubar.addMenu("File")
        
        import_action = QAction("Import...", self)
        import_action.setShortcut(QKeySequence.StandardKey.Open)
        import_action.triggered.connect(self.import_data)
        file_menu.addAction(import_action)
        
        export_action = QAction("Export...", self)
        export_action.setShortcut(QKeySequence.StandardKey.SaveAs)
        export_action.triggered.connect(self.export_data)
        file_menu.addAction(export_action)
        
        file_menu.addSeparator()
        
        exit_action = QAction("Exit", self)
        exit_action.setShortcut(QKeySequence.StandardKey.Quit)
        exit_action.triggered.connect(self.close)
        file_menu.addAction(exit_action)
        
        # Edit menu
        edit_menu = menubar.addMenu("Edit")
        
        add_action = QAction("Add Credential", self)
        add_action.setShortcut(QKeySequence.StandardKey.New)
        add_action.triggered.connect(self.add_credential)
        edit_menu.addAction(add_action)
        
        edit_action = QAction("Edit Credential", self)
        edit_action.setShortcut(QKeySequence(Qt.Key.Key_F2))
        edit_action.triggered.connect(self.edit_credential)
        edit_menu.addAction(edit_action)
        
        delete_action = QAction("Delete Credential", self)
        delete_action.setShortcut(QKeySequence.StandardKey.Delete)
        delete_action.triggered.connect(self.delete_credential)
        edit_menu.addAction(delete_action)
        
        edit_menu.addSeparator()
        
        settings_action = QAction("Settings", self)
        settings_action.setShortcut(QKeySequence.StandardKey.Preferences)
        settings_action.triggered.connect(self.show_settings)
        edit_menu.addAction(settings_action)
        
        # Tools menu
        tools_menu = menubar.addMenu("Tools")
        
        generator_action = QAction("Password Generator", self)
        generator_action.setShortcut(QKeySequence(Qt.Key.Key_F3))
        generator_action.triggered.connect(self.show_password_generator)
        tools_menu.addAction(generator_action)
        
        change_password_action = QAction("Change Master Password", self)
        change_password_action.triggered.connect(self.change_master_password)
        tools_menu.addAction(change_password_action)
        
        tools_menu.addSeparator()
        
        lock_action = QAction("Lock Application", self)
        lock_action.setShortcut(QKeySequence(Qt.Modifier.CTRL | Qt.Key.Key_L))
        lock_action.triggered.connect(self.lock_application)
        tools_menu.addAction(lock_action)
        
        # Help menu
        help_menu = menubar.addMenu("Help")
        
        about_action = QAction("About", self)
        about_action.triggered.connect(self.show_about)
        help_menu.addAction(about_action)
    
    def setup_status_bar(self):
        """Setup status bar"""
        self.status_bar = QStatusBar()
        self.setStatusBar(self.status_bar)
        
        # Add permanent widgets
        self.lock_status = QLabel("Unlocked")
        self.status_bar.addPermanentWidget(self.lock_status)
    
    def setup_shortcuts(self):
        """Setup keyboard shortcuts"""
        # Global shortcuts will be handled by menu actions
        pass
    
    def load_data(self):
        """Load credentials from database"""
        try:
            self.credentials = self.db_manager.get_all_credentials()
            self.logger.info(f"Loaded {len(self.credentials)} credentials")
            self.update_category_filter()
            self.filter_credentials()
        except Exception as e:
            self.logger.error(f"Failed to load credentials: {e}")
            QMessageBox.critical(self, "Error", f"Failed to load credentials: {e}")
    
    def update_category_filter(self):
        """Update category filter combo box"""
        # Skip if UI not initialized yet
        if not self.category_combo:
            return
            
        # Get categories from credentials, or use defaults if no credentials exist
        if self.credentials:
            categories = ["All"] + sorted(set(cred['category'] for cred in self.credentials))
        else:
            # Default categories when no credentials exist
            categories = ["All", "Email", "Shopping", "Social Media", "Banking", "Work", "Personal", "Other"]
        
        current_text = self.category_combo.currentText()
        self.category_combo.clear()
        self.category_combo.addItems(categories)
        
        # Restore selection if possible
        index = self.category_combo.findText(current_text)
        if index >= 0:
            self.category_combo.setCurrentIndex(index)
    
    def filter_credentials(self):
        """Filter credentials based on search and category"""
        # Skip if UI not initialized yet
        if not self.search_line_edit:
            return
            
        search_text = self.search_line_edit.text().lower()
        
        # Filter by search text
        if search_text:
            filtered = [
                cred for cred in self.credentials
                if (search_text in cred['title'].lower() or
                    search_text in cred['username'].lower() or
                    search_text in cred['url'].lower() or
                    search_text in cred['category'].lower())
            ]
        else:
            filtered = self.credentials.copy()
        
        # Filter by category
        if self.current_category != "All":
            filtered = [cred for cred in filtered if cred['category'] == self.current_category]
        
        self.filtered_credentials = filtered
        self.update_table()
        self.update_count_label()
    
    def filter_by_category(self, category: str):
        """Filter by selected category"""
        self.current_category = category
        self.filter_credentials()
    
    def update_table(self):
        """Update modern list-style credentials table (no avatar column)"""
        # Skip if UI not initialized yet
        if not self.table_widget:
            return
            
        self.table_widget.setRowCount(len(self.filtered_credentials))

        for row, cred in enumerate(self.filtered_credentials):
            # Title column (0)
            title_item = QTableWidgetItem(cred['title'])
            title_item.setFont(QFont("Segoe UI", 11, QFont.Weight.Bold))
            self.table_widget.setItem(row, 0, title_item)

            # Username column (1)
            username_item = QTableWidgetItem(cred['username'])
            username_item.setForeground(QColor("#1F2F9B"))
            self.table_widget.setItem(row, 1, username_item)

            # URL column (2)
            url_item = QTableWidgetItem(cred['url'])
            if cred['url']:
                url_item.setForeground(QColor("#1CA7EC"))
                url_item.setToolTip(f"Click to open: {cred['url']}")
            self.table_widget.setItem(row, 2, url_item)

            # Category column (3)
            category_item = QTableWidgetItem(cred['category'])
            category_colors = {
                'Email': '#7BD5F5',
                'Social Media': '#4ADEDE',
                'Banking': '#1CA7EC',
                'Work': '#1F2F9B',
                'General': '#666666'
            }
            color = category_colors.get(cred['category'], '#1F2F9B')
            category_item.setForeground(QColor(color))
            self.table_widget.setItem(row, 3, category_item)

            # Modified date column (4)
            try:
                from datetime import datetime
                dt = datetime.fromisoformat(cred['modified_date'])
                date_str = dt.strftime("%m/%d/%Y")
            except:
                date_str = cred['modified_date']
            date_item = QTableWidgetItem(date_str)
            date_item.setForeground(QColor("#666666"))
            self.table_widget.setItem(row, 4, date_item)

            # Store credential ID in title item for reference
            title_item.setData(Qt.ItemDataRole.UserRole, cred['id'])
    
    def update_count_label(self):
        """Update credentials count label"""
        # Skip if UI not initialized yet
        if not hasattr(self, 'count_label') or not self.count_label:
            return
            
        total = len(self.credentials)
        filtered = len(self.filtered_credentials)
        
        if filtered == total:
            self.count_label.setText(f"{total} credentials")
        else:
            self.count_label.setText(f"{filtered} of {total} credentials")
    
    def on_selection_changed(self):
        """Handle table selection change"""
        has_selection = bool(self.table_widget.selectedItems())
        
        # Enable/disable buttons (with safety checks)
        if hasattr(self, 'view_btn') and self.view_btn:
            self.view_btn.setEnabled(has_selection)
        if hasattr(self, 'edit_btn') and self.edit_btn:
            self.edit_btn.setEnabled(has_selection)
        if hasattr(self, 'delete_btn') and self.delete_btn:
            self.delete_btn.setEnabled(has_selection)
        if hasattr(self, 'copy_username_btn') and self.copy_username_btn:
            self.copy_username_btn.setEnabled(has_selection)
        if hasattr(self, 'copy_password_btn') and self.copy_password_btn:
            self.copy_password_btn.setEnabled(has_selection)
        if hasattr(self, 'open_url_btn') and self.open_url_btn:
            has_valid_url = bool(self.get_selected_credential_url()) if has_selection else False
            self.open_url_btn.setEnabled(has_valid_url)
        
        # Reset inactivity timer on user activity
        if hasattr(self, 'inactivity_timer') and self.inactivity_timer:
            self.inactivity_timer.reset()
    
    def get_selected_credential(self) -> Optional[Dict[str, Any]]:
        """Get currently selected credential"""
        current_row = self.table_widget.currentRow()
        if current_row >= 0 and current_row < len(self.filtered_credentials):
            return self.filtered_credentials[current_row]
        return None
    
    def get_selected_credential_id(self) -> Optional[int]:
        """Get selected credential ID"""
        cred = self.get_selected_credential()
        return cred['id'] if cred else None
    
    def get_selected_credential_url(self) -> Optional[str]:
        """Get selected credential URL if valid"""
        cred = self.get_selected_credential()
        if cred and cred['url']:
            url = cred['url'].strip()
            if url and (url.startswith('http://') or url.startswith('https://')):
                return url
        return None
    
    def add_credential(self):
        """Add new credential"""
        dialog = CredentialDialog(self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            data = dialog.get_data()
            
            success = self.db_manager.add_credential(
                title=data['title'],
                username=data['username'],
                password=data['password'],
                url=data['url'],
                notes=data['notes'],
                category=data['category']
            )
            
            if success:
                # Immediately save to ensure persistence
                if self.db_manager.save_database():
                    self.logger.info("Database saved after adding credential")
                else:
                    self.logger.error("Failed to save database after adding credential")
                
                self.load_data()
                self.status_bar.showMessage("Credential added successfully", 3000)
            else:
                QMessageBox.critical(self, "Error", "Failed to add credential")
    
    def edit_credential(self):
        """Edit selected credential"""
        try:
            cred_id = self.get_selected_credential_id()
            if not cred_id:
                self.logger.warning("No credential selected for editing")
                return
            
            self.logger.info(f"Attempting to edit credential ID: {cred_id}")
            
            # Get full credential with decrypted password
            credential = self.db_manager.get_credential(cred_id)
            if not credential:
                self.logger.error(f"Failed to load credential ID: {cred_id}")
                QMessageBox.critical(self, "Error", "Failed to load credential")
                return
            
            self.logger.info(f"Successfully loaded credential: {credential.get('title', 'Unknown')}")
            
            # Create and show dialog
            try:
                dialog = CredentialDialog(self, credential)
                self.logger.info("Credential dialog created successfully")
                
                if dialog.exec() == QDialog.DialogCode.Accepted:
                    self.logger.info("Dialog accepted, getting data")
                    data = dialog.get_data()
                    
                    success = self.db_manager.update_credential(
                        credential_id=cred_id,
                        title=data['title'],
                        username=data['username'],
                        password=data['password'],
                        url=data['url'],
                        notes=data['notes'],
                        category=data['category']
                    )
                    
                    if success:
                        # Immediately save to ensure persistence
                        if self.db_manager.save_database():
                            self.logger.info("Database saved after updating credential")
                        else:
                            self.logger.error("Failed to save database after updating credential")
                        
                        self.load_data()
                        self.status_bar.showMessage("Credential updated successfully", 3000)
                    else:
                        QMessageBox.critical(self, "Error", "Failed to update credential")
                        
            except Exception as dialog_error:
                self.logger.error(f"Error creating or showing credential dialog: {dialog_error}")
                QMessageBox.critical(self, "Error", f"Failed to open credential dialog: {dialog_error}")
                
        except Exception as e:
            self.logger.error(f"Unexpected error in edit_credential: {e}")
            QMessageBox.critical(self, "Error", f"An unexpected error occurred: {e}")
    
    def view_credential_details(self):
        """View credential details in a safe read-only dialog"""
        try:
            cred_id = self.get_selected_credential_id()
            if not cred_id:
                self.logger.warning("No credential selected for viewing")
                return
            
            self.logger.info(f"Attempting to view credential ID: {cred_id}")
            
            # Get full credential with decrypted password
            credential = self.db_manager.get_credential(cred_id)
            if not credential:
                self.logger.error(f"Failed to load credential ID: {cred_id}")
                QMessageBox.critical(self, "Error", "Failed to load credential")
                return
            
            self.logger.info(f"Successfully loaded credential for viewing: {credential.get('title', 'Unknown')}")
            
            # Create and show view dialog
            try:
                dialog = CredentialViewDialog(self, credential)
                self.logger.info("Credential view dialog created successfully")
                dialog.exec()
                        
            except Exception as dialog_error:
                self.logger.error(f"Error creating or showing credential view dialog: {dialog_error}")
                QMessageBox.critical(self, "Error", f"Failed to open credential view dialog: {dialog_error}")
                
        except Exception as e:
            self.logger.error(f"Unexpected error in view_credential_details: {e}")
            QMessageBox.critical(self, "Error", f"An unexpected error occurred: {e}")
    
    def delete_credential(self):
        """Delete selected credential"""
        cred = self.get_selected_credential()
        if not cred:
            return
        
        reply = QMessageBox.question(
            self, "Confirm Delete",
            f"Are you sure you want to delete '{cred['title']}'?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            success = self.db_manager.delete_credential(cred['id'])
            
            if success:
                # Immediately save to ensure persistence
                if self.db_manager.save_database():
                    self.logger.info("Database saved after deleting credential")
                else:
                    self.logger.error("Failed to save database after deleting credential")
                
                self.load_data()
                self.status_bar.showMessage("Credential deleted successfully", 3000)
            else:
                QMessageBox.critical(self, "Error", "Failed to delete credential")
    
    def copy_username(self):
        """Copy username to clipboard"""
        cred = self.get_selected_credential()
        if cred and cred['username']:
            self.clipboard_manager.copy_to_clipboard(cred['username'], auto_clear=False)
            self.status_bar.showMessage("Username copied to clipboard", 3000)
    
    def copy_password(self):
        """Copy password to clipboard"""
        cred_id = self.get_selected_credential_id()
        if not cred_id:
            return
        
        # Get credential with decrypted password
        credential = self.db_manager.get_credential(cred_id)
        if credential and credential['password']:
            auto_clear_seconds = self.settings_manager.get('clipboard_auto_clear_seconds', 30)
            self.clipboard_manager.copy_to_clipboard(credential['password'], auto_clear=True)
            self.status_bar.showMessage(
                f"Password copied to clipboard (will clear in {auto_clear_seconds}s)", 3000
            )
    
    def open_url(self):
        """Open URL in browser"""
        url = self.get_selected_credential_url()
        if url:
            import webbrowser
            try:
                webbrowser.open(url)
                self.status_bar.showMessage("URL opened in browser", 3000)
            except Exception as e:
                QMessageBox.warning(self, "Warning", f"Failed to open URL: {e}")
    
    def show_context_menu(self, position):
        """Show context menu for table"""
        if not self.table_widget.itemAt(position):
            return
        
        menu = QMenu(self)
        
        edit_action = menu.addAction("Edit")
        edit_action.triggered.connect(self.edit_credential)
        
        delete_action = menu.addAction("Delete")
        delete_action.triggered.connect(self.delete_credential)
        
        menu.addSeparator()
        
        copy_username_action = menu.addAction("Copy Username")
        copy_username_action.triggered.connect(self.copy_username)
        
        copy_password_action = menu.addAction("Copy Password")
        copy_password_action.triggered.connect(self.copy_password)
        
        url = self.get_selected_credential_url()
        if url:
            open_url_action = menu.addAction("Open URL")
            open_url_action.triggered.connect(self.open_url)
        
        menu.exec(self.table_widget.mapToGlobal(position))
    
    def show_password_generator(self):
        """Show password generator dialog"""
        dialog = PasswordGeneratorDialog(self)
        dialog.exec()
    
    def export_to_excel(self):
        """Export credentials to Excel file"""
        try:
            if not EXCEL_AVAILABLE:
                QMessageBox.critical(
                    self,
                    "Missing Dependencies",
                    "Excel export requires openpyxl.\nPlease install: pip install openpyxl"
                )
                return
                
            if not self.credentials:
                QMessageBox.information(self, "Export", "No credentials to export.")
                return
            
            # Get save file path
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Export Credentials to Excel",
                f"PasswordKeeper_Export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                "Excel Files (*.xlsx);;All Files (*)"
            )
            
            if not file_path:
                return
            
            # Create workbook and worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Credentials"
            
            # Add headers
            headers = ['Title', 'Username', 'URL', 'Category', 'Notes', 'Created', 'Modified']
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Add data (excluding passwords for security)
            for row_idx, cred in enumerate(self.credentials, 2):
                ws.cell(row=row_idx, column=1, value=cred['title'])
                ws.cell(row=row_idx, column=2, value=cred['username'])
                ws.cell(row=row_idx, column=3, value=cred['url'])
                ws.cell(row=row_idx, column=4, value=cred['category'])
                ws.cell(row=row_idx, column=5, value=cred.get('notes', ''))
                ws.cell(row=row_idx, column=6, value=cred.get('created_at', ''))
                ws.cell(row=row_idx, column=7, value=cred.get('modified_at', ''))
            
            # Save the file
            wb.save(file_path)
            
            QMessageBox.information(
                self, 
                "Export Successful", 
                f"Successfully exported {len(self.credentials)} credentials to:\n{file_path}\n\nNote: Passwords are not included for security reasons."
            )
            
        except Exception as e:
            self.logger.error(f"Export failed: {e}")
            QMessageBox.critical(self, "Export Failed", f"Failed to export credentials:\n{str(e)}")
    
    def import_from_excel(self):
        """Import credentials from Excel file"""
        try:
            if not EXCEL_AVAILABLE:
                QMessageBox.critical(
                    self,
                    "Missing Dependencies",
                    "Excel import requires openpyxl.\nPlease install: pip install openpyxl"
                )
                return
                
            # Get file path
            file_path, _ = QFileDialog.getOpenFileName(
                self,
                "Import Credentials from Excel",
                "",
                "Excel Files (*.xlsx *.xls);;All Files (*)"
            )
            
            if not file_path:
                return
            
            # Read Excel file
            wb = load_workbook(file_path)
            ws = wb.active
            
            # Get headers from first row
            headers = []
            for col in range(1, ws.max_column + 1):
                header = ws.cell(row=1, column=col).value
                if header:
                    headers.append(str(header).strip())
                else:
                    break
            
            # Validate required columns
            required_columns = ['Title', 'Username']
            missing_columns = [col for col in required_columns if col not in headers]
            
            if missing_columns:
                QMessageBox.critical(
                    self,
                    "Invalid Format",
                    f"Excel file is missing required columns: {', '.join(missing_columns)}\n\n"
                    "Required columns: Title, Username\n"
                    "Optional columns: URL, Category, Notes, Password"
                )
                return
            
            # Read data rows
            data_rows = []
            for row in range(2, ws.max_row + 1):
                row_data = {}
                for col, header in enumerate(headers, 1):
                    cell_value = ws.cell(row=row, column=col).value
                    row_data[header] = str(cell_value).strip() if cell_value else ''
                
                # Only add rows with title and username
                if row_data.get('Title') and row_data.get('Username'):
                    data_rows.append(row_data)
            
            if not data_rows:
                QMessageBox.information(self, "Import", "No valid credentials found in the Excel file.")
                return
            
            # Show preview and confirmation
            preview_text = f"Found {len(data_rows)} credentials to import:\n\n"
            for i, row in enumerate(data_rows[:5]):
                preview_text += f"• {row.get('Title', 'N/A')} ({row.get('Username', 'N/A')})\n"
            
            if len(data_rows) > 5:
                preview_text += f"... and {len(data_rows) - 5} more"
            
            reply = QMessageBox.question(
                self,
                "Confirm Import",
                preview_text + "\n\nDo you want to proceed with the import?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            
            if reply != QMessageBox.StandardButton.Yes:
                return
            
            # Import credentials
            imported_count = 0
            errors = []
            
            for i, row in enumerate(data_rows):
                try:
                    title = row.get('Title', '').strip()
                    username = row.get('Username', '').strip()
                    
                    if not title or not username:
                        errors.append(f"Row {i+2}: Missing title or username")
                        continue
                    
                    url = row.get('URL', '').strip()
                    category = row.get('Category', 'Other').strip()
                    notes = row.get('Notes', '').strip()
                    password = row.get('Password', '').strip()
                    
                    # Generate a placeholder password if none provided
                    if not password:
                        password = "ChangeMePlease123!"
                    
                    # Add credential to database
                    self.db_manager.add_credential(
                        title=title,
                        username=username,
                        password=password,
                        url=url,
                        category=category,
                        notes=notes
                    )
                    
                    imported_count += 1
                    
                except Exception as e:
                    errors.append(f"Row {i+2}: {str(e)}")
            
            # Refresh the display
            self.load_data()
            
            # Show results
            message = f"Successfully imported {imported_count} credentials."
            if errors:
                message += f"\n\nErrors encountered ({len(errors)}):\n" + "\n".join(errors[:5])
                if len(errors) > 5:
                    message += f"\n... and {len(errors) - 5} more errors"
            
            QMessageBox.information(self, "Import Complete", message)
            
        except Exception as e:
            self.logger.error(f"Import failed: {e}")
            QMessageBox.critical(self, "Import Failed", f"Failed to import credentials:\n{str(e)}")
    
    def show_settings(self):
        """Show settings dialog"""
        dialog = SettingsDialog(self, self.settings_manager)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.apply_settings()
    
    def apply_settings(self):
        """Apply current settings"""
        # Update clipboard manager
        auto_clear_seconds = self.settings_manager.get('clipboard_auto_clear_seconds', 30)
        self.clipboard_manager.set_auto_clear_seconds(auto_clear_seconds)
        
        # Update inactivity timer
        timeout_minutes = self.settings_manager.get('inactivity_timeout_minutes', 5)
        self.inactivity_timer.set_timeout(timeout_minutes)
        
        # Window size
        if self.settings_manager.get('window_remember_size', True):
            width = self.settings_manager.get('window_width', 800)
            height = self.settings_manager.get('window_height', 600)
            self.resize(width, height)
    
    def change_master_password(self):
        """Change master password"""
        from .change_password_dialog import ChangePasswordDialog
        
        dialog = ChangePasswordDialog(self, self.db_manager, self.auth_manager)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.status_bar.showMessage("Master password changed successfully", 3000)
    
    def lock_application(self):
        """Lock the application"""
        self.close()
    
    def auto_lock(self):
        """Auto-lock due to inactivity"""
        self.logger.info("Auto-locking due to inactivity")
        self.lock_application()
    
    def auto_save(self):
        """Auto-save database"""
        if self.db_manager.save_database():
            self.logger.debug("Auto-save completed")
        else:
            self.logger.warning("Auto-save failed")
    
    def import_data(self):
        """Import data from file"""
        # TODO: Implement import functionality
        QMessageBox.information(self, "Import", "Import functionality coming soon!")
    
    def export_data(self):
        """Export data to file"""
        # TODO: Implement export functionality
        QMessageBox.information(self, "Export", "Export functionality coming soon!")
    
    def show_about(self):
        """Show about dialog"""
        from core.utils import get_version
        
        about_text = f"""
        <h2>Password Keeper</h2>
        <p>Version {get_version()}</p>
        <p>A secure password manager for Windows</p>
        <p>Built with PyQt6 and advanced encryption</p>
        """
        
        QMessageBox.about(self, "About Password Keeper", about_text)
    
    def closeEvent(self, event):
        """Handle window close event"""
        # Save current window size
        if self.settings_manager.get('window_remember_size', True):
            self.settings_manager.set('window_width', self.width())
            self.settings_manager.set('window_height', self.height())
        
        # Save database
        self.db_manager.save_database()
        
        # Stop timers
        self.inactivity_timer.stop()
        self.auto_save_timer.stop()
        
        # Clear clipboard if it contains our data
        self.clipboard_manager.clear_clipboard()
        
        # Close database and clear sensitive data
        self.db_manager.close_database()
        self.auth_manager.clear_master_key()
        
        self.logger.info("Application closed")
        event.accept()
    
    def mousePressEvent(self, event):
        """Reset inactivity timer on mouse activity"""
        self.inactivity_timer.reset()
        super().mousePressEvent(event)
    
    def keyPressEvent(self, event):
        """Reset inactivity timer on keyboard activity"""
        self.inactivity_timer.reset()
        super().keyPressEvent(event)